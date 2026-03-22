"""
Meta-learning experiment script for CIFAR-100 Gibbs generalization bound experiments.

This script samples NEW class pairs from CIFAR-100 such that each class is used at most
once across all rows already present in table_meta_learning.xlsx, runs experiments for the
requested number of additional pairs, and appends one row to the Excel table after each
pair finishes.
"""

import ast
import csv
import fcntl
import random
from datetime import datetime
from pathlib import Path
from contextlib import contextmanager

import numpy as np
import torch
from openpyxl import Workbook, load_workbook

from dataset import get_cifar100_binary_dataloaders_partial_random_labels
from training import run_beta_experiments

# =========================
# Configuration flags
# =========================
TEST_MODE = True  # Set to True for quick test, False for full experiment
USE_RANDOM_LABELS = 0  # Percentage of randomly labeled data
SEEDS = [42]  # Random seeds for stability analysis
DATASET_SEED = 42  # Seed for dataset splitting/label randomization
USE_SAME_DATASET_ACROSS_SEEDS = True  # Kept for compatibility

# Core request: choose how many NEW pairs to run (classes cannot repeat across table)
NUM_NEW_PAIRS = 25
PAIR_SAMPLING_SEED = 52

# Each pair runs:
# 1) ULA for beta in [128, 1000]
# 2) GD (noise off) to approximate beta -> infinity columns
ULA_BETA_VALUES = [128, 1000]
GD_BETA_PROXY = [1000000]

# Optional side output
SAVE_SEED_SUMMARY = False

# Meta-learning table output
CREATE_NEW_TABLE_FILE = False
EXCEL_PATH = Path("table_meta_learning_CCL1W500ULA_GD1kLR001SAVAGE.xlsx")

# Fixed CIFAR-100 setup
DATASET_TYPE = "cifar100"
ALL_CIFAR100_CLASSES = list(range(100))

# Keep same learning-rate style while limiting betas to what table needs
if TEST_MODE:
    A0 = {0: 0.01, 128: 0.01, 1000: 0.01, 1000000: 0.01}
else:
    A0 = {0: 0.01, 128: 0.01, 1000: 0.01, 1000000: 0.01}

LOSS_NAME = "SAVAGE"
N_HIDDEN_LAYERS = 1
WIDTH = 500
N_TRAIN_PER_GROUP = 500
N_TEST_PER_GROUP = 100
BATCH_SIZE = 1000
SIGMA_GAUSS_PRIOR = 5.0
MIN_STEPS = 2000
MIN_STEPS_FIRST_BETA = 4000
ALPHA_AVERAGE = 0.01
ALPHA_STOP = 0.00025
ETA = 36
EPS = -1e-7

EXCEL_HEADERS = [
    "CIFAR100 Classes",
    "Mean SAVAGE training loss at beta = 128",
    "Mean SAVAGE training loss at beta = 1000",
    "Mean SAVAGE training  loss at infinity ",
    "Mean SAVAGE test  loss at infinity ",
    "Mean 01 training  loss at infinity ",
    "Mean 01 test  loss at infinity ",
]


def set_global_seed(seed: int) -> None:
    """Set random seed across torch/numpy/python for reproducibility."""
    torch.manual_seed(seed)
    np.random.seed(seed)
    random.seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False


def create_dataloaders(classes, dataset_seed):
    """Create CIFAR-100 train/test dataloaders for a given class pair/grouping."""
    return get_cifar100_binary_dataloaders_partial_random_labels(
        classes=classes,
        p=USE_RANDOM_LABELS,
        n_train_per_group=N_TRAIN_PER_GROUP,
        n_test_per_group=N_TEST_PER_GROUP,
        batch_size=BATCH_SIZE,
        random_seed=dataset_seed,
    )


def _build_meta_table_path() -> Path:
    """Build table filename using training-like convention, without timestamp."""
    if not CREATE_NEW_TABLE_FILE:
        return EXCEL_PATH

    if DATASET_TYPE == "mnist":
        dataset_prefix = "M"
    elif DATASET_TYPE in ("cifar10", "cifar100"):
        dataset_prefix = "C"
    else:
        dataset_prefix = "S"

    if USE_RANDOM_LABELS == 1:
        label_prefix = "R"
    elif USE_RANDOM_LABELS == 0:
        label_prefix = "C"
    else:
        label_prefix = "P"

    train_size = 2 * N_TRAIN_PER_GROUP
    train_size_k = f"{train_size / 1000:.0f}k"
    lr_value = A0.get(128, next(iter(A0.values()))) if isinstance(A0, dict) else A0
    lr_token = str(lr_value).replace(".", "")

    filename = (
        f"table_meta_learning_{dataset_prefix}{label_prefix}"
        f"L{N_HIDDEN_LAYERS}W{WIDTH}ULA_GD{train_size_k}"
        f"LR{lr_token}{LOSS_NAME.upper()}.xlsx"
    )
    return Path(filename)


def _build_summary_string(excel_path: Path, planned_pairs=None):
    """Build summary string for metadata appended to table workbook."""
    if planned_pairs is None:
        selected_pairs_str = "sampled online one-by-one"
    else:
        selected_pairs_str = [list(pair) for pair in planned_pairs]
    return (
        f"Meta-learning experiment summary\n"
        f"  - Table path: {excel_path}\n"
        f"  - Test mode: {TEST_MODE}\n"
        f"  - Dataset type: {DATASET_TYPE}\n"
        f"  - Random labels: {USE_RANDOM_LABELS}\n"
        f"  - Number of new pairs requested: {NUM_NEW_PAIRS}\n"
        f"  - Pair sampling seed: {PAIR_SAMPLING_SEED}\n"
        f"  - Planned pairs: {selected_pairs_str}\n"
        f"  - ULA beta values: {ULA_BETA_VALUES}\n"
        f"  - GD beta proxy: {GD_BETA_PROXY}\n"
        f"  - Seeds: {SEEDS}\n"
        f"  - Dataset seed: {DATASET_SEED}\n"
        f"  - Use same dataset across seeds: {USE_SAME_DATASET_ACROSS_SEEDS}\n"
        f"  - Loss: {LOSS_NAME}\n"
        f"  - Hidden layers: {N_HIDDEN_LAYERS}\n"
        f"  - Width: {WIDTH}\n"
        f"  - n_train_per_group: {N_TRAIN_PER_GROUP}\n"
        f"  - n_test_per_group: {N_TEST_PER_GROUP}\n"
        f"  - batch_size: {BATCH_SIZE}\n"
        f"  - a0: {A0}\n"
        f"  - sigma_gauss_prior: {SIGMA_GAUSS_PRIOR}\n"
        f"  - min_steps: {MIN_STEPS}\n"
        f"  - min_steps_first_beta: {MIN_STEPS_FIRST_BETA}\n"
        f"  - alpha_average: {ALPHA_AVERAGE}\n"
        f"  - alpha_stop: {ALPHA_STOP}\n"
        f"  - eta: {ETA}\n"
        f"  - eps: {EPS}\n"
    )


def _upsert_summary_in_excel(excel_path: Path, summary_string: str):
    """Keep exactly one summary section at the end of the workbook."""
    _ensure_workbook(excel_path)
    wb = load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]

    _remove_existing_summary(ws)

    ws.append([])
    ws.append(["Summary:", summary_string])
    wb.save(excel_path)


@contextmanager
def _excel_lock(excel_path: Path):
    """Cross-process lock for metadata table updates."""
    lock_path = Path(str(excel_path) + ".lock")
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    with open(lock_path, "w") as lock_file:
        fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX)
        try:
            yield
        finally:
            fcntl.flock(lock_file.fileno(), fcntl.LOCK_UN)


def _remove_existing_summary(ws):
    """Remove existing summary section (if present) from worksheet in-place."""

    summary_start_row = None
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == "Summary:":
            summary_start_row = row_idx
            break

    if summary_start_row is not None:
        ws.delete_rows(summary_start_row, ws.max_row - summary_start_row + 1)


def _parse_pair_cell(value):
    """Parse class pair from Excel cell like '[2,50]' -> (2, 50)."""
    if value is None:
        return None

    if isinstance(value, (list, tuple)) and len(value) == 2:
        try:
            pair = tuple(sorted((int(value[0]), int(value[1]))))
            return pair
        except (TypeError, ValueError):
            return None

    if isinstance(value, str):
        text = value.strip()
        try:
            parsed = ast.literal_eval(text)
        except (ValueError, SyntaxError):
            return None
        if isinstance(parsed, (list, tuple)) and len(parsed) == 2:
            try:
                pair = tuple(sorted((int(parsed[0]), int(parsed[1]))))
                return pair
            except (TypeError, ValueError):
                return None

    return None


def load_used_classes_from_excel(excel_path: Path):
    """Read used pairs/classes from first column to enforce class-level uniqueness."""
    used_pairs = set()
    used_classes = set()

    if not excel_path.exists():
        return used_pairs, used_classes

    wb = load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]

    for row_idx in range(2, ws.max_row + 1):
        pair = _parse_pair_cell(ws.cell(row=row_idx, column=1).value)
        if pair is None:
            continue
        used_pairs.add(pair)
        used_classes.update(pair)

    return used_pairs, used_classes


def sample_new_disjoint_pairs(num_pairs: int, used_classes: set, rng: random.Random):
    """Sample class pairs with no class reused from previous rows or within new batch."""
    if num_pairs <= 0:
        return []

    available_classes = [c for c in ALL_CIFAR100_CLASSES if c not in used_classes]
    needed_classes = 2 * num_pairs

    if needed_classes > len(available_classes):
        raise ValueError(
            f"Not enough unused CIFAR-100 classes left: need {needed_classes}, "
            f"but only {len(available_classes)} available."
        )

    rng.shuffle(available_classes)
    selected = available_classes[:needed_classes]

    pairs = []
    for i in range(0, needed_classes, 2):
        pair = tuple(sorted((selected[i], selected[i + 1])))
        pairs.append(pair)
    return pairs


def _ensure_workbook(excel_path: Path):
    """Create workbook with expected header if file does not exist yet."""
    if excel_path.exists():
        wb = load_workbook(excel_path)
        ws = wb[wb.sheetnames[0]]
        if ws.max_row == 0:
            ws.append(EXCEL_HEADERS)
            wb.save(excel_path)
        elif ws.max_row >= 1 and all(ws.cell(1, col).value is None for col in range(1, len(EXCEL_HEADERS) + 1)):
            for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            wb.save(excel_path)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(EXCEL_HEADERS)
    wb.save(excel_path)


def _read_metrics_from_csv(csv_path: Path):
    """Read beta-indexed rows from one result CSV."""
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV output not found: {csv_path}")

    rows = []
    with open(csv_path, "r", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if not row or not row.get("Beta"):
                continue
            beta_raw = row.get("Beta")
            try:
                beta_val = float(beta_raw)
            except (TypeError, ValueError):
                continue
            rows.append((beta_val, row))

    if not rows:
        raise ValueError(f"No valid data rows found in CSV: {csv_path}")

    rows.sort(key=lambda item: item[0])

    def find_row_for_beta(target_beta):
        for beta, row in rows:
            if abs(beta - target_beta) < 1e-9:
                return row
        return None

    row_128 = find_row_for_beta(128.0)
    row_1000 = find_row_for_beta(1000.0)

    return {
        "row_128": row_128,
        "row_1000": row_1000,
        "row_last": rows[-1][1],
    }


def _combine_metrics(ula_csv_path: Path, gd_csv_path: Path):
    """Build table row metrics from ULA (128/1000) and GD (infinity proxy)."""
    ula_rows = _read_metrics_from_csv(ula_csv_path)
    gd_rows = _read_metrics_from_csv(gd_csv_path)

    row_128 = ula_rows["row_128"]
    row_1000 = ula_rows["row_1000"]
    row_inf = gd_rows["row_last"]

    if row_128 is None:
        raise ValueError(f"Beta 128 not found in ULA CSV: {ula_csv_path}")
    if row_1000 is None:
        raise ValueError(f"Beta 1000 not found in ULA CSV: {ula_csv_path}")

    return {
        "train_128": float(row_128["EMA_BCE_Train"]),
        "train_1000": float(row_1000["EMA_BCE_Train"]),
        "train_inf": float(row_inf["EMA_BCE_Train"]),
        "test_inf": float(row_inf["EMA_BCE_Test"]),
        "train01_inf": float(row_inf["EMA_0-1_Train"]),
        "test01_inf": float(row_inf["EMA_0-1_Test"]),
    }


def append_result_to_excel(excel_path: Path, pair, metrics, summary_string=None):
    """Append one completed pair result row to table_meta_learning.xlsx."""
    _ensure_workbook(excel_path)
    wb = load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]

    # Keep data rows contiguous by removing summary before appending a new row.
    # Summary is re-added at the very end below (if requested).
    _remove_existing_summary(ws)

    row = [
        f"[{pair[0]},{pair[1]}]",
        float(metrics['train_128']),
        float(metrics['train_1000']),
        float(metrics['train_inf']),
        float(metrics['test_inf']),
        float(metrics['train01_inf']),
        float(metrics['test01_inf']),
    ]

    last_pair_row = 1
    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=1).value
        if _parse_pair_cell(value) is not None:
            last_pair_row = row_idx

    target_row = last_pair_row + 1
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=target_row, column=col_idx, value=value)

    wb.save(excel_path)

    if summary_string is not None:
        _upsert_summary_in_excel(excel_path, summary_string)


def _find_pair_row(ws, pair):
    for row_idx in range(2, ws.max_row + 1):
        parsed = _parse_pair_cell(ws.cell(row=row_idx, column=1).value)
        if parsed == tuple(sorted(pair)):
            return row_idx
    return None


def reserve_pair_in_excel(excel_path: Path, pair):
    """Reserve a pair row before training to avoid concurrent collisions."""
    with _excel_lock(excel_path):
        _ensure_workbook(excel_path)
        wb = load_workbook(excel_path)
        ws = wb[wb.sheetnames[0]]

        if _find_pair_row(ws, pair) is not None:
            return False

        summary_text = None
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == "Summary:":
                summary_text = ws.cell(row=row_idx, column=2).value
                break
        _remove_existing_summary(ws)

        last_pair_row = 1
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=1).value
            if _parse_pair_cell(value) is not None:
                last_pair_row = row_idx

        target_row = last_pair_row + 1
        ws.cell(row=target_row, column=1, value=f"[{pair[0]},{pair[1]}]")
        ws.cell(row=target_row, column=2, value="PENDING")

        if summary_text is not None:
            ws.append([])
            ws.append(["Summary:", summary_text])

        wb.save(excel_path)
        return True


def remove_reserved_pair_from_excel(excel_path: Path, pair):
    """Remove pair row (e.g., when a reserved pair run fails)."""
    with _excel_lock(excel_path):
        if not excel_path.exists():
            return
        wb = load_workbook(excel_path)
        ws = wb[wb.sheetnames[0]]
        row_idx = _find_pair_row(ws, pair)
        if row_idx is not None:
            ws.delete_rows(row_idx, 1)
            wb.save(excel_path)


def finalize_reserved_pair_in_excel(excel_path: Path, pair, metrics, summary_string=None):
    """Fill reserved row with metrics after successful run."""
    with _excel_lock(excel_path):
        _ensure_workbook(excel_path)
        wb = load_workbook(excel_path)
        ws = wb[wb.sheetnames[0]]

        row_idx = _find_pair_row(ws, pair)
        if row_idx is None:
            append_row = [
                f"[{pair[0]},{pair[1]}]",
                float(metrics['train_128']),
                float(metrics['train_1000']),
                float(metrics['train_inf']),
                float(metrics['test_inf']),
                float(metrics['train01_inf']),
                float(metrics['test01_inf']),
            ]
            last_pair_row = 1
            for candidate_row in range(2, ws.max_row + 1):
                value = ws.cell(row=candidate_row, column=1).value
                if _parse_pair_cell(value) is not None:
                    last_pair_row = candidate_row
            target_row = last_pair_row + 1
            for col_idx, value in enumerate(append_row, start=1):
                ws.cell(row=target_row, column=col_idx, value=value)
        else:
            ws.cell(row=row_idx, column=1, value=f"[{pair[0]},{pair[1]}]")
            ws.cell(row=row_idx, column=2, value=float(metrics['train_128']))
            ws.cell(row=row_idx, column=3, value=float(metrics['train_1000']))
            ws.cell(row=row_idx, column=4, value=float(metrics['train_inf']))
            ws.cell(row=row_idx, column=5, value=float(metrics['test_inf']))
            ws.cell(row=row_idx, column=6, value=float(metrics['train01_inf']))
            ws.cell(row=row_idx, column=7, value=float(metrics['test01_inf']))

        wb.save(excel_path)

    if summary_string is not None:
        _upsert_summary_in_excel(excel_path, summary_string)


def save_seed_stability_summary(seed_results, pair):
    """Optional: save compact stability analysis table across seeds for one pair."""
    if not seed_results:
        return None

    rows = []
    for result in seed_results:
        for csv_path in result["csv_paths"]:
            rows.append([result["seed"], result["dataset_seed"], csv_path])

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    pair_tag = f"{pair[0]}_{pair[1]}"
    summary_path = f"csv_EMA/SEED_STABILITY_CIFAR100_PAIR_{pair_tag}_{timestamp}.csv"
    with open(summary_path, "w", newline="") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(["seed", "dataset_seed", "result_csv_path"])
        writer.writerows(rows)

    return summary_path


def run_pair_experiment(pair):
    """Run ULA(128,1000) and GD for one pair; return both CSV paths."""
    print("\n" + "=" * 70)
    print(f"Running pair: {pair}")
    print(f"Dataset: CIFAR-100")
    print(f"ULA betas: {ULA_BETA_VALUES}, GD proxy beta: {GD_BETA_PROXY[-1]}")
    print("=" * 70)

    ula_seed_results = []
    gd_seed_results = []
    device = "cpu"  # change if needed

    for seed in SEEDS:
        dataset_seed = DATASET_SEED if USE_SAME_DATASET_ACROSS_SEEDS else seed

        print("\n" + "-" * 70)
        print(f"Running seed {seed} (dataset_seed={dataset_seed})")
        print("-" * 70)

        set_global_seed(seed)

        train_loader, test_loader = create_dataloaders(classes=list(pair), dataset_seed=dataset_seed)

        csv_paths_ula = run_beta_experiments(
            loss=LOSS_NAME,
            beta_values=ULA_BETA_VALUES,
            a0=A0,
            b=0.5,
            sigma_gauss_prior=SIGMA_GAUSS_PRIOR,
            device=device,
            n_hidden_layers=N_HIDDEN_LAYERS,
            width=WIDTH,
            dataset_type=DATASET_TYPE,
            use_random_labels=USE_RANDOM_LABELS,
            l_max=4.0,
            train_loader=train_loader,
            test_loader=test_loader,
            min_steps=MIN_STEPS,
            alpha_average=ALPHA_AVERAGE,
            alpha_stop=ALPHA_STOP,
            eta=ETA,
            eps=EPS,
            test_mode=TEST_MODE,
            add_grad_norm=True,
            add_noise=True,
            sgld_num=1,
            annealed=False,
            min_steps_first_beta=MIN_STEPS_FIRST_BETA,
            seed=seed,
            selected_classes=list(pair),
        )

        ula_seed_results.append(
            {
                "seed": seed,
                "dataset_seed": dataset_seed,
                "csv_paths": csv_paths_ula or [],
            }
        )

        csv_paths_gd = run_beta_experiments(
            loss=LOSS_NAME,
            beta_values=GD_BETA_PROXY,
            a0=A0,
            b=0.5,
            sigma_gauss_prior=SIGMA_GAUSS_PRIOR,
            device=device,
            n_hidden_layers=N_HIDDEN_LAYERS,
            width=WIDTH,
            dataset_type=DATASET_TYPE,
            use_random_labels=USE_RANDOM_LABELS,
            l_max=4.0,
            train_loader=train_loader,
            test_loader=test_loader,
            min_steps=MIN_STEPS,
            alpha_average=ALPHA_AVERAGE,
            alpha_stop=ALPHA_STOP,
            eta=ETA,
            eps=EPS,
            test_mode=TEST_MODE,
            add_grad_norm=True,
            add_noise=False,
            sgld_num=1,
            annealed=False,
            min_steps_first_beta=MIN_STEPS_FIRST_BETA,
            seed=seed,
            selected_classes=list(pair),
        )

        gd_seed_results.append(
            {
                "seed": seed,
                "dataset_seed": dataset_seed,
                "csv_paths": csv_paths_gd or [],
            }
        )

    if SAVE_SEED_SUMMARY:
        summary_path = save_seed_stability_summary(ula_seed_results + gd_seed_results, pair)
        if summary_path is not None:
            print(f"Seed stability summary saved to: {summary_path}")

    ula_csvs = [path for item in ula_seed_results for path in item["csv_paths"]]
    gd_csvs = [path for item in gd_seed_results for path in item["csv_paths"]]
    if not ula_csvs:
        raise RuntimeError(f"No ULA CSV output generated for pair {pair}.")
    if not gd_csvs:
        raise RuntimeError(f"No GD CSV output generated for pair {pair}.")

    return Path(ula_csvs[-1]), Path(gd_csvs[-1])


def main():
    if not SEEDS:
        raise ValueError("SEEDS must contain at least one seed value")

    excel_path = _build_meta_table_path()
    _ensure_workbook(excel_path)
    rng = random.Random(PAIR_SAMPLING_SEED)
    print(f"Table path: {excel_path}")
    print("Mode: ULA at beta 128/1000 + GD for infinity columns (online one-by-one sampling)")

    completed_pairs = 0
    attempts = 0

    while completed_pairs < NUM_NEW_PAIRS:
        attempts += 1

        used_pairs, used_classes = load_used_classes_from_excel(excel_path)
        print(f"\nCurrent table pairs: {len(used_pairs)}, used classes: {len(used_classes)}")

        try:
            pair = sample_new_disjoint_pairs(1, used_classes, rng)[0]
        except ValueError:
            print("No more class-disjoint pairs available.")
            break

        print(f">>> Reserving pair {completed_pairs + 1}/{NUM_NEW_PAIRS}: {pair}")
        if not reserve_pair_in_excel(excel_path, pair):
            print(f"Reservation race detected; pair already taken: {pair}")
            continue

        try:
            ula_csv_path, gd_csv_path = run_pair_experiment(pair)
            metrics = _combine_metrics(ula_csv_path, gd_csv_path)
            summary_string = _build_summary_string(excel_path, planned_pairs=None)
            finalize_reserved_pair_in_excel(excel_path, pair, metrics, summary_string=summary_string)
            completed_pairs += 1

            print(f"✅ Pair completed and appended to {excel_path}: {pair}")

        except Exception as exc:
            print(f"❌ Pair failed ({pair}): {exc}")
            remove_reserved_pair_from_excel(excel_path, pair)
            # Continue with next pair while preserving already written rows
            continue

    print("\n" + "=" * 70)
    print("META-LEARNING PAIR RUN COMPLETED")
    print(f"Requested pairs: {NUM_NEW_PAIRS}, completed pairs: {completed_pairs}, attempts: {attempts}")
    print("=" * 70)


if __name__ == "__main__":
    main()
