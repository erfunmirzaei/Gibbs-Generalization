"""
Recover/append meta-learning table rows from existing CSV outputs.

This script scans CSV files in `csv_EMA/`, extracts class-pair information from each
file summary, chooses one ULA file (must include beta 128 and 1000) and one GD file
for each pair, computes table metrics, and appends missing rows to the meta-learning
Excel table.
"""

import argparse
import ast
import csv
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


DEFAULT_EXCEL_PATH = Path("table_meta_learning_CCL1W500ULA_GD1kLR001SAVAGE.xlsx")
DEFAULT_CSV_DIR = Path("csv_EMA")
DEFAULT_ULA_GLOB = "CCL1W500ULA1kLR001SAVAGE*_TEST.csv"
DEFAULT_GD_GLOB = "CCL1W500GD1kLR001SAVAGE*_TEST.csv"

EXCEL_HEADERS = [
    "CIFAR100 Classes",
    "Mean SAVAGE training loss at beta = 128",
    "Mean SAVAGE training loss at beta = 1000",
    "Mean SAVAGE training  loss at infinity ",
    "Mean SAVAGE test  loss at infinity ",
    "Mean 01 training  loss at infinity ",
    "Mean 01 test  loss at infinity ",
]


def parse_args():
    parser = argparse.ArgumentParser(
        description="Recover meta-learning rows from existing ULA/GD CSV files."
    )
    parser.add_argument(
        "--excel-path",
        type=Path,
        default=DEFAULT_EXCEL_PATH,
        help="Path to target meta-learning .xlsx table.",
    )
    parser.add_argument(
        "--csv-dir",
        type=Path,
        default=DEFAULT_CSV_DIR,
        help="Directory containing experiment CSV outputs.",
    )
    parser.add_argument(
        "--ula-glob",
        type=str,
        default=DEFAULT_ULA_GLOB,
        help="Glob pattern for ULA CSV files (relative to --csv-dir).",
    )
    parser.add_argument(
        "--gd-glob",
        type=str,
        default=DEFAULT_GD_GLOB,
        help="Glob pattern for GD CSV files (relative to --csv-dir).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview rows to append without writing to Excel.",
    )
    return parser.parse_args()


def _extract_timestamp_token(path: Path):
    match = re.search(r"_(\d{8}[-_]\d{6})(?:_TEST)?\.csv$", path.name)
    if not match:
        return None
    token = match.group(1)
    fmt = "%Y%m%d-%H%M%S" if "-" in token else "%Y%m%d_%H%M%S"
    try:
        return datetime.strptime(token, fmt)
    except ValueError:
        return None


def _sort_key_for_file(path: Path):
    ts = _extract_timestamp_token(path)
    if ts is not None:
        return (ts, path.name)
    return (datetime.min, path.name)


def _ensure_workbook(excel_path: Path):
    if excel_path.exists():
        wb = load_workbook(excel_path)
        ws = wb[wb.sheetnames[0]]
        if ws.max_row == 0:
            ws.append(EXCEL_HEADERS)
            wb.save(excel_path)
        elif ws.max_row >= 1 and all(
            ws.cell(1, col).value is None for col in range(1, len(EXCEL_HEADERS) + 1)
        ):
            for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            wb.save(excel_path)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(EXCEL_HEADERS)
    wb.save(excel_path)


def _parse_pair_cell(value):
    if value is None:
        return None
    if isinstance(value, (list, tuple)) and len(value) == 2:
        try:
            return tuple(sorted((int(value[0]), int(value[1]))))
        except (TypeError, ValueError):
            return None
    if isinstance(value, str):
        text = value.strip()
        try:
            parsed = ast.literal_eval(text)
        except (ValueError, SyntaxError):
            return None
        if isinstance(parsed, (list, tuple)) and len(parsed) == 2:
            try:
                return tuple(sorted((int(parsed[0]), int(parsed[1]))))
            except (TypeError, ValueError):
                return None
    return None


def load_used_pairs_from_excel(excel_path: Path):
    used_pairs = set()
    if not excel_path.exists():
        return used_pairs

    wb = load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]
    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=1).value
        if value == "Summary:":
            break
        pair = _parse_pair_cell(value)
        if pair is None:
            continue
        used_pairs.add(pair)
    return used_pairs


def _remove_existing_summary(ws):
    summary_start_row = None
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == "Summary:":
            summary_start_row = row_idx
            break
    if summary_start_row is not None:
        ws.delete_rows(summary_start_row, ws.max_row - summary_start_row + 1)


def append_result_to_excel(excel_path: Path, pair, metrics):
    _ensure_workbook(excel_path)
    wb = load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]
    _remove_existing_summary(ws)

    row = [
        f"[{pair[0]},{pair[1]}]",
        float(metrics["train_128"]),
        float(metrics["train_1000"]),
        float(metrics["train_inf"]),
        float(metrics["test_inf"]),
        float(metrics["train01_inf"]),
        float(metrics["test01_inf"]),
    ]

    last_pair_row = 1
    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=1).value
        parsed = _parse_pair_cell(value)
        if parsed is not None:
            last_pair_row = row_idx

    target_row = last_pair_row + 1
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=target_row, column=col_idx, value=value)

    wb.save(excel_path)


def _read_csv_text(path: Path):
    with open(path, "r", newline="", encoding="utf-8") as f:
        return f.read()


def parse_pair_from_summary(path: Path):
    text = _read_csv_text(path)
    match = re.search(r"Selected classes/groups:\s*(\[[^\]]+\])", text)
    if not match:
        return None
    try:
        parsed = ast.literal_eval(match.group(1))
    except (ValueError, SyntaxError):
        return None
    if not isinstance(parsed, (list, tuple)) or len(parsed) != 2:
        return None
    try:
        return tuple(sorted((int(parsed[0]), int(parsed[1]))))
    except (TypeError, ValueError):
        return None


def read_metric_rows(path: Path):
    rows = []
    with open(path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            beta_raw = row.get("Beta") if row else None
            if not beta_raw:
                continue
            try:
                beta = float(beta_raw)
            except (TypeError, ValueError):
                continue
            rows.append((beta, row))
    rows.sort(key=lambda item: item[0])
    return rows


def find_row(rows, target_beta):
    for beta, row in rows:
        if abs(beta - target_beta) < 1e-9:
            return row
    return None


def build_best_ula_by_pair(paths):
    best = {}
    for path in sorted(paths, key=_sort_key_for_file):
        pair = parse_pair_from_summary(path)
        if pair is None:
            continue

        rows = read_metric_rows(path)
        row_128 = find_row(rows, 128.0)
        row_1000 = find_row(rows, 1000.0)
        if row_128 is None or row_1000 is None:
            continue

        best[pair] = {
            "path": path,
            "rows": rows,
            "row_128": row_128,
            "row_1000": row_1000,
        }
    return best


def build_best_gd_by_pair(paths):
    best = {}
    for path in sorted(paths, key=_sort_key_for_file):
        pair = parse_pair_from_summary(path)
        if pair is None:
            continue
        rows = read_metric_rows(path)
        if not rows:
            continue
        best[pair] = {
            "path": path,
            "rows": rows,
            "row_inf": rows[-1][1],
        }
    return best


def build_metrics(ula_info, gd_info):
    row_128 = ula_info["row_128"]
    row_1000 = ula_info["row_1000"]
    row_inf = gd_info["row_inf"]
    return {
        "train_128": float(row_128["EMA_BCE_Train"]),
        "train_1000": float(row_1000["EMA_BCE_Train"]),
        "train_inf": float(row_inf["EMA_BCE_Train"]),
        "test_inf": float(row_inf["EMA_BCE_Test"]),
        "train01_inf": float(row_inf["EMA_0-1_Train"]),
        "test01_inf": float(row_inf["EMA_0-1_Test"]),
    }


def main():
    args = parse_args()

    csv_dir = args.csv_dir
    excel_path = args.excel_path
    ula_files = list(csv_dir.glob(args.ula_glob))
    gd_files = list(csv_dir.glob(args.gd_glob))

    if not ula_files:
        raise FileNotFoundError(f"No ULA files found with glob: {csv_dir / args.ula_glob}")
    if not gd_files:
        raise FileNotFoundError(f"No GD files found with glob: {csv_dir / args.gd_glob}")

    ula_by_pair = build_best_ula_by_pair(ula_files)
    gd_by_pair = build_best_gd_by_pair(gd_files)

    candidate_pairs = sorted(set(ula_by_pair.keys()) & set(gd_by_pair.keys()))
    used_pairs = load_used_pairs_from_excel(excel_path)
    missing_pairs = [pair for pair in candidate_pairs if pair not in used_pairs]

    print(f"ULA files matched: {len(ula_files)}")
    print(f"GD files matched: {len(gd_files)}")
    print(f"Pairs with valid ULA(128,1000) and GD: {len(candidate_pairs)}")
    print(f"Pairs already in Excel: {len(used_pairs)}")
    print(f"Pairs to append: {len(missing_pairs)}")

    appended = 0
    for pair in missing_pairs:
        metrics = build_metrics(ula_by_pair[pair], gd_by_pair[pair])

        if args.dry_run:
            print(
                f"[DRY RUN] would append pair={pair} "
                f"from ULA={ula_by_pair[pair]['path'].name} GD={gd_by_pair[pair]['path'].name}"
            )
            continue

        append_result_to_excel(excel_path, pair, metrics)
        appended += 1
        print(f"✅ appended {pair}")

    print(
        f"Done. {'Would append' if args.dry_run else 'Appended'} "
        f"{len(missing_pairs) if args.dry_run else appended} rows."
    )


if __name__ == "__main__":
    main()
